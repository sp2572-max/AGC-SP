#!/usr/bin/env python3

import argparse, base64, glob, io, json, os, re, sys, time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime
from typing import Optional, Dict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font


# =========================================================================== #
# SECTION 1 — CANONICAL MASTER SCHEMA + PER-PLANT MAPPINGS
# =========================================================================== #
@dataclass(frozen=True)
class Field:
    key: str
    spanish: str
    unit: str
    dtype: str                     # float|int|str|date|time
    valid: Optional[tuple] = None  # (min,max) numeric sanity range -> auto review flag


CANONICAL_SCHEMA = [
    Field("plant",                   "Planta",                      "",     "str"),
    Field("record_date",             "Fecha",                       "",     "date"),
    Field("record_time",             "Hora (as written)",           "",     "str"),
    Field("record_time_24h",         "Hora (24h normalized)",       "",     "time"),
    Field("operator",                "Nombre",                      "",     "str"),
    Field("flow_lps",                "Caudal",                      "L/s",  "float", (0, 1000)),
    Field("turbidity_raw_ntu",       "Agua Cruda",                  "NTU",  "float", (0, 5000)),
    Field("turbidity_clarified_ntu", "Clarificada",                 "NTU",  "float", (0, 1000)),
    Field("turbidity_filtered_ntu",  "Filtrada",                    "NTU",  "float", (0, 100)),
    Field("coag_slider_pct",         "Coagulante % del deslizador", "%",    "float", (0, 100)),
    Field("coag_dose_pct_mgl",       "Coagulante Dosis segun %",    "mg/L", "float", (0, 200)),
    Field("coag_dose_probeta_mgl",   "Coagulante Dosis probeta",    "mg/L", "float", (0, 200)),
]
PROVENANCE_FIELDS = ["source_image", "extracted_at", "model", "needs_review", "review_notes"]
CANONICAL_KEYS = [f.key for f in CANONICAL_SCHEMA]
FIELD_BY_KEY = {f.key: f for f in CANONICAL_SCHEMA}
ALL_COLS = CANONICAL_KEYS + PROVENANCE_FIELDS


@dataclass
class PlantConfig:
    name: str
    columns: list                       # [(raw_spanish_header, canonical_key_or_None), ...]
    extraction_hints: str = ""
    date_format: str = "%d/%m/%y"
    bare_time_rule: str = "daytime"


SAN_JUAN_PLANES = PlantConfig(
    name="San Juan Planes",
    columns=[
        ("Fecha",                           "record_date"),
        ("Hora",                            "record_time"),
        ("Nombre",                          "operator"),
        ("Caudal (L/s)",                    "flow_lps"),
        ("Agua Cruda UTN",                  "turbidity_raw_ntu"),
        ("Clarificada UTN",                 "turbidity_clarified_ntu"),
        ("Filtrada UTN",                    "turbidity_filtered_ntu"),
        ("% del deslizador (Coagulante)",   "coag_slider_pct"),
        ("Dosis segun % mg/L (Coagulante)", "coag_dose_pct_mgl"),
        ("Dosis segun probeta mg/L (Coag)", "coag_dose_probeta_mgl"),
    ],
    extraction_hints=(
        "Handwritten water-treatment operations log. Read ONLY these columns "
        "(left group): Fecha, Hora, Nombre, Caudal, Agua Cruda UTN, Clarificada UTN, "
        "Filtrada UTN, then the three 'Coagulante' sub-columns (% del deslizador, "
        "Dosis segun % mg/L, Dosis segun probeta mg/L). IGNORE the entire 'Cloro' "
        "section and the rightmost cut-off column. Caudal is almost always 12. "
        "Dates are DD/MM/YY and often only on the first row of a day -> CARRY THE DATE "
        "DOWN until it changes. Times mix formats (7pm / 7:00 / 5am). Operators are "
        "usually Victor, Tulio, or Jose. Turbidity values are decimals (41.06, 1.79). "
        "Ignore any bleed-through / partial rows from another page at the edges."
    ),
    bare_time_rule="daytime",
)

PLANTS = {"san_juan_planes": SAN_JUAN_PLANES}


# =========================================================================== #
# SECTION 2 — NORMALIZE: raw JSON -> typed, validated canonical rows  (unchanged)
# =========================================================================== #
def _coerce(value, dtype):
    if value is None or value == "":
        return None
    try:
        if dtype == "float":
            return float(str(value).replace(",", ".").strip())
        if dtype == "int":
            return int(float(value))
        return str(value).strip()
    except (ValueError, TypeError):
        return value  # keep raw; will be flagged


def _parse_date(raw, fmt):
    if not raw:
        return None
    raw = str(raw).strip()
    for f in (fmt, "%d/%m/%Y", "%d/%m/%y", "%d-%m-%y", "%d-%m-%Y"):
        try:
            return datetime.strptime(raw, f).date()
        except ValueError:
            continue
    return None


def _to_24h(raw, bare_rule="daytime"):
    """Normalize a written time to 'HH:MM'. Honors explicit am/pm; applies
    bare_rule to bare numbers. Returns (hhmm_or_None, ok_bool)."""
    if not raw:
        return None, True
    s = str(raw).strip().lower().replace(" ", "").replace(".", "")
    m = re.match(r"^(\d{1,2})(?::(\d{2}))?(am|pm)$", s)
    if m:
        h, mins, ap = int(m.group(1)), int(m.group(2) or 0), m.group(3)
        if ap == "am":
            h = 0 if h == 12 else h
        else:
            h = h if h == 12 else h + 12
        return f"{h:02d}:{mins:02d}", True
    m = re.match(r"^(\d{1,2})(?::(\d{2}))?$", s)
    if m:
        h, mins = int(m.group(1)), int(m.group(2) or 0)
        if bare_rule == "all_am":
            h = 0 if h == 12 else h
        elif bare_rule == "all_pm":
            h = h if h == 12 else h + 12
        else:  # "daytime": 7-11 AM, 12 noon, 1-6 PM
            if 1 <= h <= 6:
                h += 12
        return f"{h:02d}:{mins:02d}", True
    return None, False  # unrecognized format -> flag for review


def normalize_extraction(raw: Dict, plant: PlantConfig, source_image: str) -> list:
    rows_out = []
    extracted_at = datetime.now().isoformat(timespec="seconds")
    model = raw.get("_model", "")
    last_date = None

    for r in raw.get("rows", []):
        out = {k: None for k in CANONICAL_KEYS}
        out["plant"] = plant.name
        review = list(r.get("uncertain_fields", []))

        for fld in FIELD_BY_KEY.values():
            if fld.key in ("plant", "record_time_24h"):
                continue
            val = r.get(fld.key)
            if fld.dtype == "date":
                parsed = _parse_date(val if val else r.get("raw_date"), plant.date_format)
                if parsed is None and (val or r.get("raw_date")):
                    review.append("record_date(unparsed)")
                parsed = parsed or last_date          # carry down
                if parsed:
                    last_date = parsed
                out[fld.key] = parsed
            else:
                coerced = _coerce(val, fld.dtype)
                out[fld.key] = coerced
                if fld.valid and isinstance(coerced, (int, float)):
                    lo, hi = fld.valid
                    if not (lo <= coerced <= hi):
                        review.append(f"{fld.key}(out_of_range)")

        t24, ok = _to_24h(out.get("record_time"), plant.bare_time_rule)
        out["record_time_24h"] = t24
        if not ok:
            review.append("record_time_24h(unparsed)")

        note = r.get("row_note", "") or ""
        out.update({
            "source_image": source_image, "extracted_at": extracted_at, "model": model,
            "needs_review": bool(review),
            "review_notes": "; ".join(filter(None, [note, ", ".join(sorted(set(review)))])),
        })
        rows_out.append(out)
    return rows_out


# =========================================================================== #
# SECTION 3 — EXTRACT: photo -> JSON via OpenRouter 
# =========================================================================== #
OPENROUTER_BASE_URL = "https://openrouter.ai/api/v1"
DEFAULT_MODEL = os.environ.get("OPENROUTER_VISION_MODEL", "nvidia/nemotron-nano-12b-v2-vl:free")
DEFAULT_WORKERS = int(os.environ.get("LOGBOOK_WORKERS", "3"))

_MIME = {".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png", ".webp": "image/webp"}


def _make_client():
    """One OpenAI-compatible client pointed at OpenRouter. Reused across photos."""
    from openai import OpenAI  # lazy: dry-run (--from-json) path needs no openai installed
    key = os.environ.get("OPENROUTER_API_KEY")
    if not key:
        raise RuntimeError(
            "OPENROUTER_API_KEY is not set. Get a key at openrouter.ai -> Keys, then:\n"
            '  export OPENROUTER_API_KEY="sk-or-..."'
        )
    return OpenAI(base_url=OPENROUTER_BASE_URL, api_key=key)


def _image_data_url(path: str) -> str:
    """Read an image and return a base64 data URL. Nemotron accepts JPEG/PNG/WEBP;
    HEIC is converted on the fly if pillow-heif is installed."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".heic":
        try:
            from PIL import Image
            import pillow_heif
            pillow_heif.register_heif_opener()
            buf = io.BytesIO()
            Image.open(path).convert("RGB").save(buf, format="JPEG", quality=92)
            return "data:image/jpeg;base64," + base64.b64encode(buf.getvalue()).decode()
        except Exception as e:
            raise RuntimeError(
                f"{os.path.basename(path)} is HEIC and couldn't be converted "
                f"(pip install pillow pillow-heif, or convert to JPG first): {e}"
            )
    mime = _MIME.get(ext)
    if not mime:
        raise RuntimeError(f"Unsupported image type '{ext}' for {os.path.basename(path)}")
    with open(path, "rb") as f:
        return f"data:{mime};base64," + base64.b64encode(f.read()).decode()


def _build_prompt(plant: PlantConfig) -> str:
    lines = []
    for raw_header, key in plant.columns:
        if key is None:
            continue
        f = FIELD_BY_KEY[key]
        rng = f", expected range {f.valid}" if f.valid else ""
        lines.append(f'  - "{key}" ({f.dtype}, unit: {f.unit or "none"}{rng}) '
                     f'— Spanish column on sheet: "{raw_header}"')
    cols = "\n".join(lines)
    return f"""You transcribe a handwritten Spanish water-treatment plant log book into JSON.

PLANT: {plant.name}

Sheet notes:
{plant.extraction_hints}

For EVERY data row in the table, output one object with these fields
(use null when a cell is blank or unreadable):
{cols}

Also include per row:
  - "raw_date": the date exactly as written, before parsing
  - "uncertain_fields": list of the field names above you are NOT confident about
  - "row_note": short note if the row is unusual, else ""

RULES
- Transcribe ONLY rows inside this table. Ignore bleed-through / partial rows at edges.
- Carry the date down to rows where it is blank but clearly the same day.
- Keep numbers exactly as written (e.g. 41.06, 0.68). Never round or invent.
- Ambiguous digit: give your best read AND list that field in "uncertain_fields".
- Do NOT include any explanation or reasoning. Output ONLY a JSON object:
  {{"rows": [ {{...}} ], "page_note": "..."}}
"""


def _parse_json(text: str) -> Dict:
    text = (text or "").strip()
    # some reasoning models prepend a <think>...</think> trace — drop it
    if "</think>" in text:
        text = text.split("</think>")[-1].strip()
    if text.startswith("```"):
        text = text.split("```")[1]
        if text.lstrip().lower().startswith("json"):
            text = text.lstrip()[4:]
        text = text.strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        s, e = text.find("{"), text.rfind("}")
        if s != -1 and e != -1:
            return json.loads(text[s:e + 1])
        raise


# transient conditions worth retrying with backoff (rate limits / provider hiccups)
_RETRYABLE = ("429", "rate", "timeout", "timed out", "502", "503", "overloaded", "temporarily")


def extract_photo(path: str, plant: PlantConfig, model: str = DEFAULT_MODEL,
                  client=None, max_retries: int = 5) -> Dict:
    if client is None:
        client = _make_client()
    data_url = _image_data_url(path)
    messages = [{
        "role": "user",
        "content": [
            {"type": "text", "text": _build_prompt(plant)},
            {"type": "image_url", "image_url": {"url": data_url}},
        ],
    }]
    last_err = None
    for attempt in range(max_retries):
        try:
            resp = client.chat.completions.create(
                model=model,
                messages=messages,
                temperature=0,
                # turn off Nemotron's reasoning trace -> cleaner JSON, faster/cheaper
                extra_body={"reasoning": {"enabled": False}},
            )
            data = _parse_json(resp.choices[0].message.content)
            data["_model"] = f"openrouter:{model}"
            return data
        except Exception as e:  # noqa: BLE001 — we classify below
            last_err = e
            if any(tok in str(e).lower() for tok in _RETRYABLE) and attempt < max_retries - 1:
                time.sleep(min(2 ** attempt, 30))  # 1,2,4,8,16,30s...
                continue
            raise
    raise RuntimeError(f"failed after {max_retries} retries: {last_err}")


# =========================================================================== #
# SECTION 4 — WRITE: append to the plant's Excel file with RED review cells  (unchanged)
# =========================================================================== #
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RED_FONT = Font(color="9C0006", bold=True)


def _load_existing(path, plant_name):
    if not os.path.exists(path):
        return pd.DataFrame(columns=ALL_COLS)
    try:
        return pd.read_excel(path, sheet_name=plant_name[:31]).reindex(columns=ALL_COLS)
    except Exception:
        return pd.DataFrame(columns=ALL_COLS)


def _schema_map_df():
    recs = []
    for plant in PLANTS.values():
        for raw_header, canon in plant.columns:
            fld = FIELD_BY_KEY.get(canon)
            recs.append({"plant": plant.name, "raw_column_spanish": raw_header,
                         "canonical_field": canon,
                         "unit": fld.unit if fld else "", "dtype": fld.dtype if fld else ""})
    return pd.DataFrame(recs)


def _flagged_fields(notes):
    return {k for k in CANONICAL_KEYS if isinstance(notes, str) and k in notes}


def write_workbook(df, path, plant_name):
    df = df.reindex(columns=ALL_COLS)
    if not df.empty:
        df = df.sort_values(["record_date", "record_time_24h"],
                            na_position="last").reset_index(drop=True)
    review = df[df["needs_review"] == True].copy()  # noqa: E712
    sheet = plant_name[:31] or "PLANT"

    with pd.ExcelWriter(path, engine="openpyxl") as xl:
        df.to_excel(xl, sheet_name=sheet, index=False)
        review.to_excel(xl, sheet_name="REVIEW_QUEUE", index=False)
        _schema_map_df().to_excel(xl, sheet_name="SCHEMA_MAP", index=False)

    wb = load_workbook(path)
    col_idx = {c: i + 1 for i, c in enumerate(ALL_COLS)}
    for ws_name, data in ((sheet, df), ("REVIEW_QUEUE", review)):
        ws = wb[ws_name]
        for r, (_, row) in enumerate(data.reset_index(drop=True).iterrows(), start=2):
            if not bool(row.get("needs_review")):
                continue
            nr = ws.cell(row=r, column=col_idx["needs_review"]); nr.fill = RED; nr.font = RED_FONT
            for key in _flagged_fields(row.get("review_notes")):
                ws.cell(row=r, column=col_idx[key]).fill = RED
    wb.save(path)
    return df, review


# =========================================================================== #
# SECTION 5 — CLI
# =========================================================================== #
def _rows_from_json(plant, json_globs):
    rows = []
    for pattern in json_globs:
        for jf in sorted(glob.glob(pattern)):
            with open(jf, encoding="utf-8") as f:
                raw = json.load(f)
            rows.extend(normalize_extraction(raw, plant, raw.get("_source_image", os.path.basename(jf))))
    return rows


def _rows_from_photos(plant, photos_dir, done, model=DEFAULT_MODEL, workers=DEFAULT_WORKERS):
    all_imgs = [p for p in sorted(glob.glob(os.path.join(photos_dir, "*")))
                if p.lower().endswith((".jpg", ".jpeg", ".png", ".webp", ".heic"))]
    todo = [p for p in all_imgs if os.path.basename(p) not in done]
    skipped = len(all_imgs) - len(todo)
    total = len(todo)

    if skipped:
        print(f"[skip] {skipped} photo(s) already in the file", file=sys.stderr)
    if total == 0:
        print("Nothing new to process.", file=sys.stderr)
        return []

    client = _make_client()          # built once, shared across worker threads
    workers = max(1, min(workers, total))
    print(f"Processing {total} photo(s) with {workers} worker(s) via {model} ...", file=sys.stderr)

    def work(img):
        name = os.path.basename(img)
        t0 = time.time()
        try:
            extracted = extract_photo(img, plant, model=model, client=client)
            new_rows = normalize_extraction(extracted, plant, name)
            return name, new_rows, time.time() - t0, None
        except Exception as e:  # noqa: BLE001
            return name, None, time.time() - t0, e

    rows, ok_count, fail_count = [], 0, 0
    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = [pool.submit(work, img) for img in todo]
        for i, fut in enumerate(as_completed(futures), start=1):
            name, new_rows, dt, err = fut.result()
            if err is None:
                rows.extend(new_rows)
                ok_count += 1
                flagged = sum(1 for r in new_rows if r["needs_review"])
                print(f"[{i}/{total}] {name} done in {dt:0.0f}s -> "
                      f"{len(new_rows)} rows ({flagged} flagged)", file=sys.stderr)
            else:
                fail_count += 1
                print(f"[{i}/{total}] {name} FAILED after {dt:0.0f}s -> {err}", file=sys.stderr)

    print(f"Done: {ok_count} succeeded, {fail_count} failed, "
          f"{len(rows)} total rows extracted.", file=sys.stderr)
    return rows


def main():
    ap = argparse.ArgumentParser(description="Water log-book photos -> Excel (OpenRouter).")
    ap.add_argument("--plant", required=True, choices=list(PLANTS))
    ap.add_argument("--photos", help="folder of photos (live extraction)")
    ap.add_argument("--from-json", nargs="+", help="glob(s) of pre-extracted JSON (dry run)")
    ap.add_argument("--out", default="plant.xlsx")
    ap.add_argument("--model", default=DEFAULT_MODEL,
                    help=f"OpenRouter model slug (default: {DEFAULT_MODEL}). "
                         f"Try google/gemini-2.5-flash for tougher handwriting.")
    ap.add_argument("--workers", type=int, default=DEFAULT_WORKERS,
                    help=f"parallel requests (default: {DEFAULT_WORKERS}; keep low on the free tier)")
    ap.add_argument("--reprocess", action="store_true", help="re-extract even if already in file")
    args = ap.parse_args()
    plant = PLANTS[args.plant]

    existing = _load_existing(args.out, plant.name)
    done = set() if args.reprocess else set(existing["source_image"].dropna().unique())

    if args.from_json:
        new = [r for r in _rows_from_json(plant, args.from_json)
               if args.reprocess or r["source_image"] not in done]
    elif args.photos:
        new = _rows_from_photos(plant, args.photos, done, model=args.model, workers=args.workers)
    else:
        ap.error("provide --photos (live) or --from-json (dry run)")

    new_df = pd.DataFrame(new, columns=ALL_COLS)
    if args.reprocess and not new_df.empty:
        existing = existing[~existing["source_image"].isin(new_df["source_image"].unique())]
    combined = pd.concat([existing, new_df], ignore_index=True).drop_duplicates(
        subset=["source_image", "record_date", "record_time", "operator"], keep="last")

    df, review = write_workbook(combined, args.out, plant.name)
    print(f"Appended {len(new_df)} new rows. Total {len(df)} in '{args.out}' "
          f"({len(review)} flagged red for review).")


if __name__ == "__main__":
    main()
